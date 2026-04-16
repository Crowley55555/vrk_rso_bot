from __future__ import annotations

import asyncio
import logging
import uuid
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from api.config import (
    ACCIDENTS_SHEET,
    COMPLETED_SHEET,
    IN_PROGRESS_SHEET,
    LOG_SHEET,
    NOT_STARTED_SHEET,
)
from api.services.sqlite_service import SQLiteService


TASK_HEADERS = ["Дата", "Наименование", "Комментарий", "Ответственные", "Срок", "Кто добавил", "_uuid"]
ACCIDENTS_HEADERS = [
    "Дата и время",
    "Краткое описание",
    "Подробное описание",
    "Ответственные",
    "Срочность",
    "Кто сообщил",
    "_uuid",
]
LOG_HEADERS = ["Дата и время", "Кто", "Действие", "Задача", "Лист", "Подробности", "_uuid"]
UUID_COLUMN_LETTER = "G"

SUPPORTED_SHEETS = [
    NOT_STARTED_SHEET,
    IN_PROGRESS_SHEET,
    COMPLETED_SHEET,
    ACCIDENTS_SHEET,
    LOG_SHEET,
]


logger = logging.getLogger(__name__)


class ExcelServiceError(Exception):
    """Ошибка при работе с xlsx-файлом."""


class ExcelService:
    """Сервис синхронизации данных между SQLite и xlsx."""

    def __init__(self, excel_path: Path) -> None:
        self._excel_path = excel_path

    def export_sheet(self, sheet_name: str, db_rows: list[dict[str, Any]]) -> None:
        """Полностью перезаписывает лист данными из БД."""
        self._ensure_parent_dir()
        workbook = self._load_or_create_workbook()
        # Удаляем лист и создаём заново — это единственный надёжный способ
        # полной очистки в openpyxl без риска дублирования строк.
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        worksheet = workbook.create_sheet(sheet_name)

        worksheet.append(self._headers_for_sheet(sheet_name))
        for row in db_rows:
            worksheet.append(
                [
                    row.get("col_a", ""),
                    row.get("col_b", ""),
                    row.get("col_c", ""),
                    row.get("col_d", ""),
                    row.get("col_e", ""),
                    row.get("col_f", ""),
                    row.get("row_uuid", ""),
                ]
            )

        # Служебный uuid всегда в колонке G и всегда скрыт.
        worksheet.column_dimensions[UUID_COLUMN_LETTER].hidden = True
        workbook.save(self._excel_path)
        workbook.close()

    async def import_sheet(
        self,
        sheet_name: str,
        sqlite_service: SQLiteService,
        disk_mtime: float | None = None,
        preloaded_rows: list[dict[str, Any]] | None = None,
        preloaded_read_stats: dict[str, int] | None = None,
    ) -> dict[str, Any]:
        """Импортирует лист из xlsx в SQLite с merge по uuid и возвращает отчёт."""
        effective_mtime = disk_mtime if disk_mtime is not None else self._local_file_mtime()
        if preloaded_rows is not None and preloaded_read_stats is not None:
            disk_rows = preloaded_rows
            read_stats = preloaded_read_stats
        else:
            disk_rows, read_stats = await asyncio.to_thread(self._read_sheet_rows, sheet_name)
        db_rows = await sqlite_service.get_all(sheet_name)

        db_uuids = {str(row["row_uuid"]) for row in db_rows}
        disk_rows, duplicate_uuid_regenerated = self._make_disk_rows_unique(disk_rows)

        db_by_uuid = {str(row["row_uuid"]): row for row in db_rows}
        disk_by_uuid = {str(row["row_uuid"]): row for row in disk_rows}
        disk_uuids = set(disk_by_uuid.keys())

        last_upload_raw = await sqlite_service.get_sync_meta("last_upload_at")
        last_upload_at = self._safe_float(last_upload_raw, default=0.0)

        inserted_count = 0
        deleted_count = 0
        updated_count = 0
        skipped_db_newer_count = 0
        skipped_unchanged_count = 0

        # 1. Только на диске -> INSERT в SQLite.
        for row_uuid in sorted(disk_uuids - db_uuids):
            disk_row = disk_by_uuid[row_uuid]
            await sqlite_service.insert(
                sheet_name,
                disk_row["row_data"],
                row_uuid=row_uuid,
                created_at=effective_mtime,
                updated_at=effective_mtime,
                row_order=disk_row.get("row_order", 0),
            )
            inserted_count += 1

        # 2. Только в БД -> удаляем только то, что уже точно было выгружено на диск.
        for row_uuid in sorted(db_uuids - disk_uuids):
            db_row = db_by_uuid[row_uuid]
            # Удаляем строку только если она точно была выгружена на диск ранее.
            # Если upload ни разу не выполнялся (last_upload_at == 0) —
            # не удаляем ничего, чтобы не потерять данные добавленные через бота.
            if last_upload_at == 0:
                skipped_db_newer_count += 1
                continue
            created_at = self._safe_float(db_row.get("created_at"), default=0.0)
            if created_at > last_upload_at:
                skipped_db_newer_count += 1
                continue
            await sqlite_service.delete_by_id(sheet_name, int(db_row["id"]))
            deleted_count += 1

        # 3. Есть и в БД и на диске -> если диск свежее, обновляем SQLite.
        if effective_mtime > 0:
            for row_uuid in sorted(db_uuids & disk_uuids):
                db_row = db_by_uuid[row_uuid]
                db_updated_at = self._safe_float(db_row.get("updated_at"), default=0.0)
                if effective_mtime <= db_updated_at:
                    skipped_db_newer_count += 1
                    continue

                row_id = int(db_row["id"])
                disk_row_order = int(disk_by_uuid[row_uuid].get("row_order", 0))
                current_row_order = int(db_row.get("row_order") or 0)
                disk_values = self._normalize_row_values(disk_by_uuid[row_uuid]["row_data"])
                db_values = self._normalize_row_values(
                    [
                        db_row.get("col_a", ""),
                        db_row.get("col_b", ""),
                        db_row.get("col_c", ""),
                        db_row.get("col_d", ""),
                        db_row.get("col_e", ""),
                        db_row.get("col_f", ""),
                    ]
                )
                if disk_values == db_values and current_row_order == disk_row_order:
                    skipped_unchanged_count += 1
                    continue
                if disk_values != db_values:
                    await sqlite_service.update_row(sheet_name, row_id, disk_values)
                if current_row_order != disk_row_order:
                    await sqlite_service.update_row_order(sheet_name, row_id, disk_row_order)
                updated_count += 1

        skipped_count = read_stats["empty_rows_skipped"] + skipped_db_newer_count + skipped_unchanged_count
        report = {
            "sheet_name": sheet_name,
            "read_count": read_stats["scanned_rows"],
            "imported_count": inserted_count + updated_count,
            "skipped_count": skipped_count,
            "skipped_reasons": {
                "empty_a_to_f": read_stats["empty_rows_skipped"],
                "db_newer_or_equal": skipped_db_newer_count,
                "unchanged_row": skipped_unchanged_count,
            },
            "details": {
                "inserted": inserted_count,
                "updated": updated_count,
                "deleted": deleted_count,
                "uuid_generated": read_stats["uuid_generated"],
                "duplicate_uuid_regenerated": duplicate_uuid_regenerated,
            },
        }
        logger.info(
            "Импорт листа %s: прочитано=%s импортировано=%s пропущено=%s причины=%s",
            sheet_name,
            report["read_count"],
            report["imported_count"],
            report["skipped_count"],
            report["skipped_reasons"],
        )
        return report

    async def read_sheet_for_import(self, sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, int]]:
        """Считывает лист из xlsx для последующего merge без изменения БД."""
        return await asyncio.to_thread(self._read_sheet_rows, sheet_name)

    async def export_all_sheets(self, sqlite_service: SQLiteService) -> None:
        """Выгружает все поддерживаемые листы из SQLite в xlsx."""
        rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in SUPPORTED_SHEETS:
            rows_by_sheet[sheet_name] = await sqlite_service.get_all(sheet_name)

        await asyncio.to_thread(self._export_all_sheets_sync, rows_by_sheet)

    def _export_all_sheets_sync(self, rows_by_sheet: dict[str, list[dict[str, Any]]]) -> None:
        """Синхронная часть полной выгрузки всех листов."""
        self._ensure_parent_dir()
        workbook = self._load_or_create_workbook()

        for sheet_name in SUPPORTED_SHEETS:
            # Удаляем лист и создаём заново — это единственный надёжный способ
            # полной очистки в openpyxl без риска дублирования строк.
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            worksheet = workbook.create_sheet(sheet_name)

            worksheet.append(self._headers_for_sheet(sheet_name))
            for row in rows_by_sheet.get(sheet_name, []):
                worksheet.append(
                    [
                        row.get("col_a", ""),
                        row.get("col_b", ""),
                        row.get("col_c", ""),
                        row.get("col_d", ""),
                        row.get("col_e", ""),
                        row.get("col_f", ""),
                        row.get("row_uuid", ""),
                    ]
                )
            worksheet.column_dimensions[UUID_COLUMN_LETTER].hidden = True

        workbook.save(self._excel_path)
        workbook.close()

    def _read_sheet_rows(self, sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, int]]:
        """Читает строки листа и возвращает данные в формате merge."""
        workbook = self._load_or_create_workbook()
        worksheet = self._ensure_sheet(workbook, sheet_name)

        rows: list[dict[str, Any]] = []
        scanned_rows = 0
        empty_rows_skipped = 0
        uuid_generated = 0
        max_row = worksheet.max_row

        for row_idx in range(2, max_row + 1):
            scanned_rows += 1
            # Читаем строго A-F для данных и G для служебного uuid.
            row_data = self._extract_row_data(sheet_name, row_idx, worksheet)
            row_uuid = self._to_text(worksheet.cell(row=row_idx, column=7).value).strip()

            # Пропускаем только действительно пустые строки по A-F.
            if not any(value.strip() for value in row_data):
                empty_rows_skipped += 1
                continue
            if not row_uuid:
                row_uuid = str(uuid.uuid4())
                uuid_generated += 1
            rows.append(
                {
                    "row_uuid": row_uuid,
                    "row_data": row_data,
                    "row_order": len(rows) + 1,
                }
            )

        workbook.close()
        return rows, {
            "scanned_rows": scanned_rows,
            "empty_rows_skipped": empty_rows_skipped,
            "uuid_generated": uuid_generated,
        }

    def _extract_row_data(self, sheet_name: str, row_idx: int, worksheet: Any) -> list[str]:
        """Возвращает данные строки строго по колонкам A-F с учётом типа листа."""
        values = [self._to_text(worksheet.cell(row=row_idx, column=col).value) for col in range(1, 7)]
        if sheet_name in {NOT_STARTED_SHEET, IN_PROGRESS_SHEET, COMPLETED_SHEET}:
            # Листы задач: A дата, B название, C комментарий, D ответственные, E срок, F кто добавил.
            return values
        if sheet_name == ACCIDENTS_SHEET:
            # Лист аварий: A дата, B краткое описание, C подробности, D ответственные, E срочность, F кто добавил.
            return values
        if sheet_name == LOG_SHEET:
            # Лист лога: A дата и время, B кто, C действие, D задача, E лист, F подробности.
            return values
        raise ExcelServiceError(f"Неизвестный лист для импорта: {sheet_name}")

    @staticmethod
    def _make_disk_rows_unique(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
        """Убирает дубли uuid внутри одного листа, не теряя строки."""
        seen: set[str] = set()
        regenerated = 0
        normalized: list[dict[str, Any]] = []

        for row in rows:
            current_uuid = str(row["row_uuid"])
            while current_uuid in seen:
                current_uuid = str(uuid.uuid4())
                regenerated += 1
            seen.add(current_uuid)
            normalized.append(
                {
                    "row_uuid": current_uuid,
                    "row_data": row["row_data"],
                    "row_order": row.get("row_order", 0),
                }
            )
        return normalized, regenerated

    @staticmethod
    def _normalize_row_values(values: list[Any]) -> list[str]:
        """Нормализует значения строки для корректного сравнения A-F."""
        normalized = ["" if value is None else str(value) for value in values[:6]]
        if len(normalized) < 6:
            normalized.extend([""] * (6 - len(normalized)))
        return normalized

    def _load_or_create_workbook(self) -> Workbook:
        if self._excel_path.exists():
            return load_workbook(self._excel_path)

        workbook = Workbook()
        # Удаляем дефолтный лист, чтобы создать строго нужную структуру.
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for sheet_name in SUPPORTED_SHEETS:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(self._headers_for_sheet(sheet_name))
            worksheet.column_dimensions[UUID_COLUMN_LETTER].hidden = True
        workbook.save(self._excel_path)
        return workbook

    @staticmethod
    def _safe_float(value: Any, default: float) -> float:
        try:
            if value is None:
                return default
            return float(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _to_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    def _local_file_mtime(self) -> float:
        """Возвращает mtime локального xlsx для сравнения свежести."""
        if not self._excel_path.exists():
            return 0.0
        return float(self._excel_path.stat().st_mtime)

    def _ensure_parent_dir(self) -> None:
        self._excel_path.parent.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _headers_for_sheet(sheet_name: str) -> list[str]:
        if sheet_name == ACCIDENTS_SHEET:
            return ACCIDENTS_HEADERS.copy()
        if sheet_name == LOG_SHEET:
            return LOG_HEADERS.copy()
        return TASK_HEADERS.copy()

    @staticmethod
    def _ensure_sheet(workbook: Workbook, sheet_name: str):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        return workbook.create_sheet(sheet_name)
